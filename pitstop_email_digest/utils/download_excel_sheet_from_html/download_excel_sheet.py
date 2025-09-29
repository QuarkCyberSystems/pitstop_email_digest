from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import Workbook
import frappe

def html_table_to_excel(html_string, excel_file_path):
    soup = BeautifulSoup(html_string, "html.parser")
    tables = soup.find_all("table")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for t_idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet(title=f"Sheet{t_idx}")
        rows = table.find_all("tr")
        row_idx = 1

        # Track merged cells
        merged_cells = set()

        for row in rows:
            col_idx = 1
            for cell in row.find_all(["td", "th"]):
                # Skip columns that are inside merged cells
                while (row_idx, col_idx) in merged_cells:
                    col_idx += 1

                value = cell.get_text(strip=True)
                ws.cell(row=row_idx, column=col_idx, value=value)

                # Handle colspan and rowspan
                colspan = int(cell.get("colspan", 1))
                rowspan = int(cell.get("rowspan", 1))

                if colspan > 1 or rowspan > 1:
                    end_row = row_idx + rowspan - 1
                    end_col = col_idx + colspan - 1
                    ws.merge_cells(start_row=row_idx, start_column=col_idx,
                                   end_row=end_row, end_column=end_col)

                    # Mark all cells in the merged area except top-left
                    for r in range(row_idx, end_row + 1):
                        for c in range(col_idx, end_col + 1):
                            if (r, c) != (row_idx, col_idx):
                                merged_cells.add((r, c))

                col_idx += colspan
            row_idx += 1
    
     # Instead of saving to disk, write to in-memory buffer
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)

    # Send to Frappe response
    frappe.response["filename"] = excel_file_path
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"